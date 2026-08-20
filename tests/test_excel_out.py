"""Tests for the styled .xlsx writer."""
import datetime as dt

import openpyxl
import pandas as pd
import pytest

from monoamine_calc.config import load_defaults
from monoamine_calc.engine import run
from monoamine_calc.excel_out import default_filename, write_excel

STD_COLUMNS = ["STD_ID", "MHPG", "NE", "DHBA", "DOPAC", "DA", "5-HIAA", "5-HT"]
SAMPLE_COLUMNS = ["SAMPLE_ID", "MHPG", "NE", "DHBA", "DOPAC", "DA", "5-HIAA", "5-HT", "PROTEIN", "VOLUME"]


@pytest.fixture
def result():
    config = load_defaults()
    standards = pd.DataFrame(
        [["S1", 100, 200, 500, 300, 400, 250, 350], ["S2", 120, 180, 520, 280, 420, 270, 330]],
        columns=STD_COLUMNS,
    )
    samples = pd.DataFrame(
        [["A1", 90, 210, 480, 310, 390, 240, 360, 1.5, 0.1]],
        columns=SAMPLE_COLUMNS,
    )
    return run(samples, standards, config), config


def test_default_filename_format():
    name = default_filename(dt.date(2026, 3, 7))
    assert name == "monoamine concentration 07.03.26.xlsx"


def test_write_excel_creates_expected_sheets(tmp_path, result):
    calc_result, config = result
    out = write_excel(calc_result, config, tmp_path / "out.xlsx")
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    assert "concentration_wide" in wb.sheetnames
    assert "std_chain" in wb.sheetnames
    assert "constants_used" in wb.sheetnames


def test_write_excel_formats_and_freezes_header(tmp_path, result):
    calc_result, config = result
    out = write_excel(calc_result, config, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["concentration_wide"]
    assert ws.freeze_panes == "A2"
    # SAMPLE_ID is column A (text), the second column onward are numeric ng/mg.
    assert ws["B2"].number_format == "0.000"
    assert ws["A1"].value == "SAMPLE_ID"


def test_write_excel_includes_warnings_sheet_only_when_present(tmp_path):
    config = load_defaults()
    standards = pd.DataFrame([["S1", 100, 200, 500, 300, 400, 250, 350]], columns=STD_COLUMNS)
    zero_samples = pd.DataFrame(
        [["A1", 90, 210, 0, 310, 390, 240, 360, 0, 0.1]], columns=SAMPLE_COLUMNS
    )
    calc_result = run(zero_samples, standards, config)
    assert calc_result.warnings  # sanity check on the fixture

    out = write_excel(calc_result, config, tmp_path / "warn.xlsx")
    wb = openpyxl.load_workbook(out)
    assert "warnings" in wb.sheetnames
