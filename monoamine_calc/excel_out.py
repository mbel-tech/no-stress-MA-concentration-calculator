"""Styled .xlsx writer, ported from the R script's openxlsx section.

Same filename pattern ("monoamine concentration DD.MM.YY.xlsx"), 0.000
number format on numeric columns, auto column widths, frozen header row.

Two extra sheets versus the R script, so an exported file is
self-documenting:
- "std_chain": the computed A -> F standard dilution table
- "constants_used": a flat dump of the Config that produced this file

Uses openpyxl directly (already a dependency of pandas' Excel writer) so no
extra package is required beyond what's already installed.
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import Config
from .engine import CalculationResult

NUMBER_FORMAT = "0.000"


def default_filename(today: _dt.date | None = None) -> str:
    today = today or _dt.date.today()
    return f"monoamine concentration {today.strftime('%d.%m.%y')}.xlsx"


def _write_dataframe(ws: Worksheet, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in df.itertuples(index=False):
        ws.append(list(row))

    numeric_col_idx = [
        i + 1 for i, dtype in enumerate(df.dtypes) if pd.api.types.is_numeric_dtype(dtype)
    ]
    for col_idx in numeric_col_idx:
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, df.shape[0] + 2):
            ws[f"{col_letter}{row_idx}"].number_format = NUMBER_FORMAT

    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(f"{v:.3f}" if isinstance(v, float) else str(v)) for v in df[col]]
        )
        ws.column_dimensions[get_column_letter(i)].width = max(8, min(40, max_len + 2))

    ws.freeze_panes = "A2"


def _constants_rows(config: Config) -> pd.DataFrame:
    rows = []
    for a in config.analytes:
        rows.append({
            "Analyte": a.name,
            "Standard weight (mg)": a.standard_weight_mg,
            "MW ratio": a.mw_ratio,
            "MW correction applied": a.apply_mw_correction,
            "Formula": a.formula,
            "Internal standard": a.is_internal_standard,
        })
    df = pd.DataFrame(rows)
    d = config.dilution
    meta = pd.DataFrame([
        {"Analyte": "-- dilution chain --", "Standard weight (mg)": None, "MW ratio": None,
         "MW correction applied": None, "Formula": None, "Internal standard": None},
        {"Analyte": "Step A divisor", "Standard weight (mg)": d.step_a_divisor},
        {"Analyte": "Step B multiplier", "Standard weight (mg)": d.step_b_multiplier},
        {"Analyte": "Step C multiplier", "Standard weight (mg)": d.step_c_multiplier},
        {"Analyte": "Step D divisor", "Standard weight (mg)": d.step_d_divisor},
        {"Analyte": "Step E divisor", "Standard weight (mg)": d.step_e_divisor},
        {"Analyte": "Step F divisor", "Standard weight (mg)": d.step_f_divisor},
        {"Analyte": "STD step used", "Standard weight (mg)": config.std_step},
        {"Analyte": "Rounding (decimals)", "Standard weight (mg)": config.round_decimals},
    ])
    return pd.concat([df, meta], ignore_index=True)


def write_excel(result: CalculationResult, config: Config, out_path: Path) -> Path:
    out_path = Path(out_path)
    wb = Workbook()

    ws_main = wb.active
    ws_main.title = "concentration_wide"
    _write_dataframe(ws_main, result.concentrations)

    ws_chain = wb.create_sheet("std_chain")
    _write_dataframe(ws_chain, result.std_chain)

    ws_const = wb.create_sheet("constants_used")
    _write_dataframe(ws_const, _constants_rows(config))

    if result.warnings:
        ws_warn = wb.create_sheet("warnings")
        _write_dataframe(ws_warn, pd.DataFrame({"warning": result.warnings}))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
